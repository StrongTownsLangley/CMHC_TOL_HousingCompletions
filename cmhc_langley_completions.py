"""
CMHC Housing Completions - Langley (DM)
========================================
Downloads and extracts all available completions data:
  - Annual files for 2010-2023
  - Monthly CSD files for 2024-2025 (cumulative to get annual totals)

Outputs:
  - langley_dm_completions.csv                (annual summary, all years)
  - langley_dm_monthly_detail_2024_2025.csv   (month-by-month 2024-2025)
  - langley_dm_completions.txt                (formatted summary table)

All downloaded files are cached in ./cmhc_cache/ so re-runs are fast.

Usage:  python cmhc_langley_completions.py
Requires: pip install openpyxl requests
"""

import os
import requests
import openpyxl
import csv
from pathlib import Path

# ── Configuration ──────────────────────────────────────────────────

ASSET_ROOT = "https://assets.cmhc-schl.gc.ca"

MONTHLY_BASE = (
    f"{ASSET_ROOT}/sites/cmhc/professional/"
    "housing-markets-data-and-research/housing-data-tables/"
    "housing-market-data/housing-information-monthly"
)

MONTHS = {
    1: "january", 2: "february", 3: "march", 4: "april",
    5: "may", 6: "june", 7: "july", 8: "august",
    9: "september", 10: "october", 11: "november", 12: "december"
}

CACHE_DIR = Path("cmhc_cache")
OUTPUT_CSV = "langley_dm_completions.csv"
OUTPUT_MONTHLY_CSV = "langley_dm_monthly_detail_2024_2025.csv"
OUTPUT_TXT = "langley_dm_completions.txt"


# ── URL builders ───────────────────────────────────────────────────

def annual_urls(year):
    """
    CMHC has changed the download path at least six times:
      2022-2023  sites/cmhc/professional/.../{name}-en.xlsx
      2020-2021  sites/cmhc/professional/.../{name}.xlsx
      2019       sites/cmhc/data-research/data-tables/.../{name}.xlsx
      2018       sf/project/cmhc/xls/data-tables/.../{name}.xlsx
      2016-2017  sf/project/cmhc/pubsandreports/excel/scs-2-1-{name}.xlsx
      2010-2015  sf/project/cmhc/pubsandreports/excel/table_2_1_{year}_e.xlsx
    """
    name = f"housing-completions-dwelling-type-{year}"
    hcd = "housing-completions-dwelling-type"
    return [
        f"{ASSET_ROOT}/sites/cmhc/professional/housing-markets-data-and-research/housing-data-tables/housing-market-data/{hcd}/{name}-en.xlsx",
        f"{ASSET_ROOT}/sites/cmhc/professional/housing-markets-data-and-research/housing-data-tables/housing-market-data/{hcd}/{name}.xlsx",
        f"{ASSET_ROOT}/sites/cmhc/data-research/data-tables/{hcd}/{name}.xlsx",
        f"{ASSET_ROOT}/sf/project/cmhc/xls/data-tables/{hcd}/{name}.xlsx",
        f"{ASSET_ROOT}/sf/project/cmhc/pubsandreports/excel/scs-2-1-{name}.xlsx",
        f"{ASSET_ROOT}/sf/project/cmhc/pubsandreports/excel/table_2_1_{year}_e.xlsx",
    ]


def monthly_url(year, month):
    yy = str(year)[2:]
    return (
        f"{MONTHLY_BASE}/{year}/{MONTHS[month]}/"
        f"starts-completions-under-construction-{month:02d}-{yy}-en.xlsx"
    )


# ── Helpers ────────────────────────────────────────────────────────

def download(url, filepath, quiet=False):
    try:
        r = requests.get(url, timeout=30)
        if r.status_code == 200:
            filepath.write_bytes(r.content)
            return True
        if not quiet:
            print(f"  HTTP {r.status_code}")
        return False
    except Exception as e:
        if not quiet:
            print(f"  Error: {e}")
        return False


def clean(v):
    """Cell value to int. Handles None, '--', '-', comma-formatted strings."""
    if v is None or v == "-" or v == "--":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        try:
            return int(v.replace(",", "").strip())
        except ValueError:
            return 0
    return 0


def find_cached(urls):
    """Return cached filepath if any URL's filename exists in CACHE_DIR."""
    for url in urls:
        candidate = CACHE_DIR / url.split("/")[-1].split("?")[0]
        if candidate.exists():
            return candidate
    return None


def download_first(urls):
    """Try each URL until one succeeds. Return filepath or None."""
    for i, url in enumerate(urls):
        filename = url.split("/")[-1].split("?")[0]
        filepath = CACHE_DIR / filename
        if download(url, filepath, quiet=(i < len(urls) - 1)):
            return filepath
    return None


# ── Extractors ─────────────────────────────────────────────────────

def extract_annual(filepath):
    """
    Extract Langley (DM) from an annual completions file.
    Returns dict {single, semi, row, apt, total} or None.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  Could not open: {e}")
        return None

    ws = None
    for name in ["CSD", "CSD - SDR"]:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = list(row)
        csd_idx = None
        for i, v in enumerate(vals):
            if v and ("Langley (DM)" in str(v) or "Langley DM" in str(v)):
                csd_idx = i
                break
        if csd_idx is None:
            continue

        num = vals[csd_idx + 1:]
        if len(num) >= 5:
            return {
                "single": clean(num[0]),
                "semi": clean(num[1]),
                "row": clean(num[2]),
                "apt": clean(num[3]),
                "total": clean(num[4]),
            }

    print(f"  Langley DM not found")
    return None


def extract_monthly(filepath):
    """
    Extract Langley DM from Table H10 (BC) in a monthly CSD file.
    Returns dict {monthly: {...}, cumulative: {...}, under_construction} or None.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  Could not open: {e}")
        return None

    if "Table H10" not in wb.sheetnames:
        print(f"  No Table H10")
        return None

    ws = wb["Table H10"]

    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = list(row)
        if not (vals[1] and "Langley DM" in str(vals[1])):
            continue
        return {
            "monthly": {
                "single": clean(vals[12]), "semi": clean(vals[13]),
                "row": clean(vals[14]), "apt": clean(vals[15]),
                "total": clean(vals[16]),
            },
            "cumulative": {
                "single": clean(vals[17]), "semi": clean(vals[18]),
                "row": clean(vals[19]), "apt": clean(vals[20]),
                "total": clean(vals[21]),
            },
            "under_construction": clean(vals[26]),
        }

    print(f"  Langley DM not found")
    return None


# ── Main ───────────────────────────────────────────────────────────

def main():
    CACHE_DIR.mkdir(exist_ok=True)

    annual_results = []
    monthly_results = []

    # ── Phase 1: Annual files (2010-2023) ──────────────────────────

    print("=" * 60)
    print("Phase 1: Annual completions (2010-2023)")
    print("=" * 60)

    for year in range(2010, 2024):
        urls = annual_urls(year)
        print(f"  {year}: ", end="")

        filepath = find_cached(urls)
        if filepath:
            print("(cached) ", end="")
        else:
            filepath = download_first(urls)
            if not filepath:
                print("SKIPPED (all URLs failed)")
                continue

        data = extract_annual(filepath)
        if data:
            annual_results.append({"year": year, **data})
            print(
                f"S:{data['single']:>4}  "
                f"Sm:{data['semi']:>4}  "
                f"R:{data['row']:>4}  "
                f"A:{data['apt']:>5}  "
                f"T:{data['total']:>5}"
            )
        else:
            print("EXTRACT FAILED")

    # ── Phase 2: Monthly files (2024-2025) ─────────────────────────

    print()
    print("=" * 60)
    print("Phase 2: Monthly CSD data (2024-2025)")
    print("=" * 60)

    for year in [2024, 2025]:
        for month in range(1, 13):
            url = monthly_url(year, month)
            filename = url.split("/")[-1]
            filepath = CACHE_DIR / filename

            print(f"  {year}-{month:02d}: ", end="")

            if filepath.exists():
                print("(cached) ", end="")
            else:
                if not download(url, filepath):
                    print("not available")
                    continue

            data = extract_monthly(filepath)
            if data:
                monthly_results.append({
                    "year": year,
                    "month": month,
                    "month_name": MONTHS[month],
                    **{f"monthly_{k}": v for k, v in data["monthly"].items()},
                    **{f"cumulative_{k}": v for k, v in data["cumulative"].items()},
                    "under_construction": data["under_construction"],
                })
                m, c = data["monthly"], data["cumulative"]
                print(
                    f"month:{m['total']:>5}  "
                    f"YTD:{c['total']:>5}  "
                    f"u/c:{data['under_construction']:>5}"
                )
            else:
                print("EXTRACT FAILED")

    # ── Derive annual totals for 2024/2025 from last available month

    for year in [2024, 2025]:
        year_rows = [r for r in monthly_results if r["year"] == year]
        if not year_rows:
            continue
        last = year_rows[-1]
        annual_results.append({
            "year": year,
            "single": last["cumulative_single"],
            "semi": last["cumulative_semi"],
            "row": last["cumulative_row"],
            "apt": last["cumulative_apt"],
            "total": last["cumulative_total"],
            "_through": last["month_name"],
            "_under_construction": last["under_construction"],
        })

    annual_results.sort(key=lambda r: r["year"])

    # ── Write CSVs ─────────────────────────────────────────────────

    csv_fields = ["year", "single", "semi", "row", "apt", "total"]
    with open(OUTPUT_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(annual_results)
    print(f"\nSaved: {OUTPUT_CSV}")

    if monthly_results:
        mfields = list(monthly_results[0].keys())
        with open(OUTPUT_MONTHLY_CSV, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=mfields)
            writer.writeheader()
            writer.writerows(monthly_results)
        print(f"Saved: {OUTPUT_MONTHLY_CSV}")

    # ── Build summary ──────────────────────────────────────────────

    lines = []
    lines.append("=" * 68)
    lines.append("HOUSING COMPLETIONS - TOWNSHIP OF LANGLEY (DM)")
    lines.append("Source: CMHC Starts and Completions Survey")
    lines.append("=" * 68)
    lines.append("")

    hdr = f"{'Year':>6}  {'Single':>8}  {'Semi':>6}  {'Row':>6}  {'Apt':>7}  {'Total':>7}  Note"
    lines.append(hdr)
    lines.append("-" * 68)

    for r in annual_results:
        note = ""
        if "_through" in r:
            if r["_through"] != "december":
                note = f"(through {r['_through']})"
        lines.append(
            f"{r['year']:>6}  "
            f"{r['single']:>8}  "
            f"{r['semi']:>6}  "
            f"{r['row']:>6}  "
            f"{r['apt']:>7}  "
            f"{r['total']:>7}  {note}"
        )

    lines.append("-" * 68)

    # Averages (full years only)
    full_years = [
        r for r in annual_results
        if "_through" not in r or r.get("_through") == "december"
    ]

    if len(full_years) >= 5:
        last5 = full_years[-5:]
        avg5 = sum(r["total"] for r in last5) / 5
        span = f"{last5[0]['year']}-{last5[-1]['year']}"
        lines.append(
            f"{'':>6}  {'':>8}  {'':>6}  {'':>6}  {'':>7}  "
            f"{avg5:>7.0f}  5-yr avg ({span})"
        )

    avg_all = sum(r["total"] for r in full_years) / max(len(full_years), 1)
    lines.append(
        f"{'':>6}  {'':>8}  {'':>6}  {'':>6}  {'':>7}  "
        f"{avg_all:>7.0f}  avg all full years ({len(full_years)} yrs)"
    )

    # Under construction
    if monthly_results:
        last_m = monthly_results[-1]
        lines.append("")
        lines.append(
            f"Units under construction as of "
            f"{last_m['month_name'].title()} {last_m['year']}: "
            f"{last_m['under_construction']:,}"
        )

    lines.append("")
    lines.append("Notes:")
    lines.append("  'Completions' = move-in ready units per CMHC (all proposed")
    lines.append("  construction work performed, or up to 10% remaining).")
    lines.append("  '--' or missing values treated as 0.")
    lines.append("  2010-2023: annual 'Housing Completions by Dwelling Type' tables.")
    lines.append("  2024-2025: derived from cumulative monthly CSD data.")

    summary = "\n".join(lines)
    print(f"\n{summary}")

    with open(OUTPUT_TXT, "w") as f:
        f.write(summary + "\n")
    print(f"\nSaved: {OUTPUT_TXT}")


if __name__ == "__main__":
    main()
