"""
CMHC Housing Completions by Municipality
=========================================
Downloads CMHC completions data and extracts results for any municipality.

  Annual files:  2010-2023 (Housing Completions by Dwelling Type)
  Monthly files: 2024-2025 (Starts, Completions and Under Construction by CSD)

Examples:
  python cmhc_annual.py                          # Langley Township, all years
  python cmhc_annual.py --start 2015 --end 2023  # Langley Township, 2015-2023
  python cmhc_annual.py -m "Bur*"                # Burnaby, Burns Lake, etc.
  python cmhc_annual.py -m "Sur*" --start 2020   # Surrey from 2020
  python cmhc_annual.py --list                   # print all available municipalities
  python cmhc_annual.py --refresh                # re-download all files

Requires: pip install openpyxl requests
"""

import argparse
import fnmatch
import os
import requests
import openpyxl
import csv
import sys
from pathlib import Path

# ── Configuration ──────────────────────────────────────────────────

ASSET_ROOT = "https://assets.cmhc-schl.gc.ca"

MONTHLY_BASE = (
    f"{ASSET_ROOT}/sites/cmhc/professional/"
    "housing-markets-data-and-research/housing-data-tables/"
    "housing-market-data/housing-information-monthly"
)

MONTHS = {
    1: "january", 2: "february", 3: "march", 4: "april",
    5: "may", 6: "june", 7: "july", 8: "august",
    9: "september", 10: "october", 11: "november", 12: "december"
}

CACHE_DIR = Path("cmhc_cache")

DEFAULT_MATCH = "Langley (DM)"
ANNUAL_START = 2010
ANNUAL_END = 2023
MONTHLY_YEARS = [2024, 2025]


# ── URL builders ───────────────────────────────────────────────────

def annual_urls(year):
    name = f"housing-completions-dwelling-type-{year}"
    hcd = "housing-completions-dwelling-type"
    return [
        f"{ASSET_ROOT}/sites/cmhc/professional/housing-markets-data-and-research/housing-data-tables/housing-market-data/{hcd}/{name}-en.xlsx",
        f"{ASSET_ROOT}/sites/cmhc/professional/housing-markets-data-and-research/housing-data-tables/housing-market-data/{hcd}/{name}.xlsx",
        f"{ASSET_ROOT}/sites/cmhc/data-research/data-tables/{hcd}/{name}.xlsx",
        f"{ASSET_ROOT}/sf/project/cmhc/xls/data-tables/{hcd}/{name}.xlsx",
        f"{ASSET_ROOT}/sf/project/cmhc/pubsandreports/excel/scs-2-1-{name}.xlsx",
        f"{ASSET_ROOT}/sf/project/cmhc/pubsandreports/excel/table_2_1_{year}_e.xlsx",
    ]


def monthly_url(year, month):
    yy = str(year)[2:]
    return (
        f"{MONTHLY_BASE}/{year}/{MONTHS[month]}/"
        f"starts-completions-under-construction-{month:02d}-{yy}-en.xlsx"
    )


# ── Helpers ────────────────────────────────────────────────────────

def download(url, filepath, quiet=False):
    try:
        r = requests.get(url, timeout=30)
        if r.status_code == 200:
            filepath.write_bytes(r.content)
            return True
        if not quiet:
            print(f"  HTTP {r.status_code}")
        return False
    except Exception as e:
        if not quiet:
            print(f"  Error: {e}")
        return False


def clean(v):
    if v is None or v == "-" or v == "--":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        try:
            return int(v.replace(",", "").strip())
        except ValueError:
            return 0
    return 0


def find_cached(urls):
    for url in urls:
        candidate = CACHE_DIR / url.split("/")[-1].split("?")[0]
        if candidate.exists():
            return candidate
    return None


def download_first(urls):
    for i, url in enumerate(urls):
        filename = url.split("/")[-1].split("?")[0]
        filepath = CACHE_DIR / filename
        if download(url, filepath, quiet=(i < len(urls) - 1)):
            return filepath
    return None


def matches(csd_name, pattern):
    """Case-insensitive fnmatch. Pattern without wildcards matches as substring."""
    name = str(csd_name)
    if "*" in pattern or "?" in pattern:
        return fnmatch.fnmatch(name.lower(), pattern.lower())
    return pattern.lower() in name.lower()


def ensure_file(urls, refresh=False):
    """Return filepath from cache or download. Respect --refresh."""
    if not refresh:
        cached = find_cached(urls)
        if cached:
            return cached, True
    return download_first(urls), False


def ensure_monthly_file(year, month, refresh=False):
    url = monthly_url(year, month)
    filename = url.split("/")[-1]
    filepath = CACHE_DIR / filename
    if not refresh and filepath.exists():
        return filepath, True
    if download(url, filepath):
        return filepath, False
    return None, False


# ── Open a workbook (annual) ───────────────────────────────────────

def open_annual_sheet(filepath):
    """Open an annual file and return the CSD worksheet."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  Could not open: {e}")
        return None

    for name in ["CSD", "CSD - SDR"]:
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]


# ── Extract matching rows ──────────────────────────────────────────

def extract_annual_all(filepath):
    """
    Return list of (csd_name, {single, semi, row, apt, total}) for every CSD.
    """
    ws = open_annual_sheet(filepath)
    if ws is None:
        return []

    results = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = list(row)
        # Find the CSD name column (contains parenthetical type codes like DM, CY, T)
        csd_idx = None
        for i, v in enumerate(vals):
            s = str(v) if v else ""
            if "(" in s and ")" in s and i >= 2:
                csd_idx = i
                break
        if csd_idx is None:
            continue

        csd_name = str(vals[csd_idx]).strip()
        num = vals[csd_idx + 1:]
        if len(num) >= 5:
            results.append((csd_name, {
                "single": clean(num[0]),
                "semi": clean(num[1]),
                "row": clean(num[2]),
                "apt": clean(num[3]),
                "total": clean(num[4]),
            }))
    return results


def extract_monthly_all(filepath):
    """
    Return list of (csd_name, {monthly:{...}, cumulative:{...}, under_construction})
    for every CSD across all provincial tables (H1-H10).
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  Could not open: {e}")
        return []

    results = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Table H"):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            vals = list(row)
            if len(vals) < 27:
                continue
            csd_raw = vals[1]
            if not csd_raw or not isinstance(csd_raw, str):
                continue
            # Skip header/total rows
            if csd_raw.startswith("Total") or csd_raw.startswith("Starts"):
                continue
            # CSD names in monthly files lack parentheses: "Langley DM" not "Langley (DM)"
            # Check if it looks like a data row (has numbers after position 2)
            if clean(vals[6]) == 0 and clean(vals[16]) == 0 and clean(vals[26]) == 0:
                if vals[6] != 0 and vals[16] != 0:
                    continue

            csd_name = csd_raw.strip()
            results.append((csd_name, {
                "monthly": {
                    "single": clean(vals[12]), "semi": clean(vals[13]),
                    "row": clean(vals[14]), "apt": clean(vals[15]),
                    "total": clean(vals[16]),
                },
                "cumulative": {
                    "single": clean(vals[17]), "semi": clean(vals[18]),
                    "row": clean(vals[19]), "apt": clean(vals[20]),
                    "total": clean(vals[21]),
                },
                "under_construction": clean(vals[26]),
            }))
    return results


def normalize_name(name):
    """
    Normalize CSD names so annual and monthly formats match.
    Annual: 'Langley (DM)'  Monthly: 'Langley DM'
    """
    return name.replace("(", "").replace(")", "").strip()


# ── List command ───────────────────────────────────────────────────

def do_list(refresh):
    """Download the most recent annual file and print all CSD names."""
    CACHE_DIR.mkdir(exist_ok=True)

    # Use 2023 as reference year
    urls = annual_urls(2023)
    filepath, cached = ensure_file(urls, refresh)
    if not filepath:
        print("Could not download reference file.")
        return

    rows = extract_annual_all(filepath)
    names = sorted(set(name for name, _ in rows))
    print(f"{len(names)} municipalities in 2023 data:\n")
    for n in names:
        print(f"  {n}")


# ── Main extraction ────────────────────────────────────────────────

def do_extract(pattern, start, end, refresh, output_dir):
    CACHE_DIR.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)

    annual_results = {}   # {csd_name: [{year, single, semi, row, apt, total}, ...]}
    monthly_results = {}  # {csd_name: [{year, month, month_name, ...}, ...]}

    # Determine year ranges
    annual_years = [y for y in range(ANNUAL_START, ANNUAL_END + 1) if start <= y <= end]
    monthly_years = [y for y in MONTHLY_YEARS if start <= y <= end]

    # ── Phase 1: Annual ────────────────────────────────────────────

    if annual_years:
        print("=" * 60)
        print(f"Phase 1: Annual completions ({annual_years[0]}-{annual_years[-1]})")
        print(f"Pattern: {pattern}")
        print("=" * 60)

        for year in annual_years:
            urls = annual_urls(year)
            print(f"  {year}: ", end="")

            filepath, cached = ensure_file(urls, refresh)
            if cached:
                print("(cached) ", end="")
            if not filepath:
                print("SKIPPED")
                continue

            rows = extract_annual_all(filepath)
            matched = [(n, d) for n, d in rows if matches(n, pattern)]

            if not matched:
                print(f"no match for '{pattern}'")
                continue

            for csd_name, data in matched:
                annual_results.setdefault(csd_name, [])
                annual_results[csd_name].append({"year": year, **data})

            names = ", ".join(n for n, _ in matched)
            totals = " + ".join(str(d["total"]) for _, d in matched)
            print(f"{names} -> {totals}")

    # ── Phase 2: Monthly ───────────────────────────────────────────

    if monthly_years:
        print()
        print("=" * 60)
        print(f"Phase 2: Monthly CSD data ({monthly_years[0]}-{monthly_years[-1]})")
        print("=" * 60)

        for year in monthly_years:
            for month in range(1, 13):
                print(f"  {year}-{month:02d}: ", end="")

                filepath, cached = ensure_monthly_file(year, month, refresh)
                if cached:
                    print("(cached) ", end="")
                if not filepath:
                    print("not available")
                    continue

                rows = extract_monthly_all(filepath)
                # Match using normalized names
                matched = [
                    (n, d) for n, d in rows
                    if matches(n, normalize_name(pattern))
                    or matches(f"{n.rsplit(' ', 1)[0]} ({n.rsplit(' ', 1)[-1]})" if " " in n else n, pattern)
                ]

                if not matched:
                    print(f"no match")
                    continue

                for csd_name, data in matched:
                    monthly_results.setdefault(csd_name, [])
                    monthly_results[csd_name].append({
                        "year": year,
                        "month": month,
                        "month_name": MONTHS[month],
                        **{f"monthly_{k}": v for k, v in data["monthly"].items()},
                        **{f"cumulative_{k}": v for k, v in data["cumulative"].items()},
                        "under_construction": data["under_construction"],
                    })

                totals = ", ".join(
                    f"{n}:{d['cumulative']['total']}" for n, d in matched
                )
                print(f"YTD: {totals}")

    # ── Combine: derive annual from monthly ────────────────────────

    # Build a unified name mapping (monthly names lack parentheses)
    all_annual_names = set(annual_results.keys())
    all_monthly_names = set(monthly_results.keys())

    # For each monthly CSD, try to find its annual equivalent
    name_map = {}  # monthly_name -> display_name
    for mname in all_monthly_names:
        # Try to find matching annual name
        matched_annual = [
            aname for aname in all_annual_names
            if normalize_name(aname) == normalize_name(mname)
        ]
        if matched_annual:
            name_map[mname] = matched_annual[0]
        else:
            name_map[mname] = mname

    # Merge monthly into annual using last available month
    for mname, rows in monthly_results.items():
        display = name_map.get(mname, mname)
        for year in monthly_years:
            year_rows = [r for r in rows if r["year"] == year]
            if not year_rows:
                continue
            last = year_rows[-1]
            annual_results.setdefault(display, [])
            annual_results[display].append({
                "year": year,
                "single": last["cumulative_single"],
                "semi": last["cumulative_semi"],
                "row": last["cumulative_row"],
                "apt": last["cumulative_apt"],
                "total": last["cumulative_total"],
                "_through": last["month_name"],
                "_under_construction": last["under_construction"],
            })

    # Sort each CSD's results by year
    for name in annual_results:
        annual_results[name].sort(key=lambda r: r["year"])

    if not annual_results:
        print(f"\nNo data found for pattern '{pattern}'.")
        return

    # ── Output: one CSV + TXT per municipality ──────────────────────

    saved = []

    for csd_name, rows in sorted(annual_results.items()):
        safe = (
            csd_name
            .replace("(", "").replace(")", "")
            .replace(" ", "_").replace(".", "")
            .strip("_").lower()
        )
        csv_file = output_dir / f"completions_{safe}.csv"
        txt_file = output_dir / f"completions_{safe}.txt"

        # CSV
        with open(csv_file, "w", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["year", "single", "semi", "row", "apt", "total"],
                extrasaction="ignore",
            )
            writer.writeheader()
            for r in rows:
                writer.writerow(r)

        # TXT
        lines = []
        lines.append("=" * 68)
        lines.append(f"HOUSING COMPLETIONS - {csd_name.upper()}")
        lines.append("Source: CMHC Starts and Completions Survey")
        lines.append("=" * 68)
        lines.append("")

        hdr = f"{'Year':>6}  {'Single':>8}  {'Semi':>6}  {'Row':>6}  {'Apt':>7}  {'Total':>7}  Note"
        lines.append(hdr)
        lines.append("-" * 68)

        for r in rows:
            note = ""
            if "_through" in r and r["_through"] != "december":
                note = f"(through {r['_through']})"
            lines.append(
                f"{r['year']:>6}  "
                f"{r['single']:>8}  "
                f"{r['semi']:>6}  "
                f"{r['row']:>6}  "
                f"{r['apt']:>7}  "
                f"{r['total']:>7}  {note}"
            )

        lines.append("-" * 68)

        full = [r for r in rows if "_through" not in r or r.get("_through") == "december"]
        if len(full) >= 5:
            last5 = full[-5:]
            avg5 = sum(r["total"] for r in last5) / 5
            span = f"{last5[0]['year']}-{last5[-1]['year']}"
            lines.append(f"{'':>6}  {'':>8}  {'':>6}  {'':>6}  {'':>7}  {avg5:>7.0f}  5-yr avg ({span})")
        if full:
            avg_all = sum(r["total"] for r in full) / len(full)
            lines.append(f"{'':>6}  {'':>8}  {'':>6}  {'':>6}  {'':>7}  {avg_all:>7.0f}  avg ({len(full)} full yrs)")

        # Under construction from most recent monthly
        display_monthly = monthly_results.get(csd_name, [])
        if not display_monthly:
            for mname, dname in name_map.items():
                if dname == csd_name:
                    display_monthly = monthly_results.get(mname, [])
                    break
        if display_monthly:
            last_m = display_monthly[-1]
            lines.append(
                f"Under construction ({last_m['month_name'].title()} "
                f"{last_m['year']}): {last_m['under_construction']:,}"
            )

        lines.append("")
        lines.append("Notes:")
        lines.append("  'Completions' = move-in ready units per CMHC (all proposed")
        lines.append("  construction work performed, or up to 10% remaining).")
        lines.append("  2010-2023: annual 'Housing Completions by Dwelling Type' tables.")
        lines.append("  2024-2025: derived from cumulative monthly CSD data.")

        summary = "\n".join(lines)
        print(f"\n{summary}")

        with open(txt_file, "w") as f:
            f.write(summary + "\n")

        saved.append((csd_name, csv_file, txt_file))

    # Print file list
    print()
    print("=" * 60)
    print(f"Generated {len(saved)} file(s):")
    for csd_name, csv_file, txt_file in saved:
        print(f"  {csd_name}")
        print(f"    {csv_file}")
        print(f"    {txt_file}")


# ── CLI ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="CMHC housing completions by municipality.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""examples:
  %(prog)s                          # Langley Township, all years
  %(prog)s --start 2015 --end 2023  # Langley Township, 2015-2023
  %(prog)s -m "Bur*"               # Burnaby, Burns Lake, etc.
  %(prog)s -m "Sur*" --start 2020  # Surrey from 2020
  %(prog)s -o results              # output to ./results/ instead of ./output/
  %(prog)s --list                  # print all municipality names
  %(prog)s --refresh               # re-download all files""",
    )
    parser.add_argument(
        "-m", "--match",
        default=DEFAULT_MATCH,
        help="Municipality pattern. Supports * and ? wildcards. "
             "Without wildcards, matches as substring. "
             f"Default: '{DEFAULT_MATCH}'",
    )
    parser.add_argument("-o", "--output", default="output", help="Output directory (default: output)")
    parser.add_argument("--start", type=int, default=ANNUAL_START, help=f"Start year (default: {ANNUAL_START})")
    parser.add_argument("--end", type=int, default=MONTHLY_YEARS[-1], help=f"End year (default: {MONTHLY_YEARS[-1]})")
    parser.add_argument("--list", action="store_true", help="List all available municipality names and exit")
    parser.add_argument("--refresh", action="store_true", help="Re-download files even if cached")

    args = parser.parse_args()

    if args.list:
        do_list(args.refresh)
    else:
        do_extract(args.match, args.start, args.end, args.refresh, Path(args.output))


if __name__ == "__main__":
    main()
