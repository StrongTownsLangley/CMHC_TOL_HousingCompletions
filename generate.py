#!/usr/bin/env python3
"""
One-time generator: build all CSVs and mapping.csv.

Reads cmhc_cache/ and pop_cache/ (which must already exist from
prior runs of cmhc_annual.py and pop_annual.py), finds every BC
municipality that appears in both datasets, then calls the two
scripts to produce every CSV. Finally writes mapping.csv for the
web frontend.

Usage:
    python generate.py
    python generate.py --pop-cache pop_cache --cmhc-cache cmhc_cache
    python generate.py -o output
"""

import argparse
import csv
import os
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required.  pip install openpyxl")


# ── Constants ──────────────────────────────────────────────────────

AREA_TYPE_LABELS = {
    "CY": "City",
    "DM": "District Municipality",
    "T": "Town",
    "VL": "Village",
    "IM": "Island Municipality",
    "RGM": "Regional Municipality",
    "IGD": "Indian Government District",
}

CHART_PREFIX = {
    "CY": "City",
    "DM": "District",
    "T": "Town",
    "VL": "Village",
    "IM": "Municipality",
    "RGM": "Municipality",
}


# ── Helpers ────────────────────────────────────────────────────────

def slugify(name):
    out = []
    for ch in name.lower():
        if ch.isalnum():
            out.append(ch)
        elif out and out[-1] != "_":
            out.append("_")
    return "".join(out).strip("_")


def pop_base_name(pop_name):
    for suffix in (", City of", ", District Municipality", ", Town of", ", Village of"):
        if pop_name.endswith(suffix):
            return pop_name[: -len(suffix)].strip()
    return pop_name


def display_name(base, area_type):
    label = AREA_TYPE_LABELS.get(area_type, "")
    return f"{base} ({label})" if label else base


def chart_name(base, area_type):
    prefix = CHART_PREFIX.get(area_type, "")
    return f"{prefix} of {base}" if prefix else base


# ── Read BC Stats population names ────────────────────────────────

def read_pop_names(cache_dir):
    candidates = [
        "pop_municipal_subprov_areas.xlsx",
        "pop_municipal_subprov_areas_2001_2011.xlsx",
    ]
    found = None
    for fn in candidates:
        path = os.path.join(cache_dir, fn)
        if os.path.isfile(path):
            found = path
            break
    if found is None:
        return None

    wb = openpyxl.load_workbook(found, data_only=True, read_only=True)
    if "Mun Name Sort" not in wb.sheetnames:
        wb.close()
        return None

    ws = wb["Mun Name Sort"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "Name":
            header_idx = i
            break
    if header_idx is None:
        return None

    results = []
    seen = set()
    for row in rows[header_idx + 1:]:
        name = row[0]
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        area_type = row[1].strip() if row[1] else ""
        if not area_type or area_type not in AREA_TYPE_LABELS:
            continue
        key = (name, area_type)
        if key in seen:
            continue
        seen.add(key)
        results.append((name, area_type))
    return results


# ── Read CMHC CSD names for BC ────────────────────────────────────

def read_cmhc_bc_names(cache_dir):
    bc_labels = {"B.C./C.-B.", "B.C.", "British Columbia"}
    non_bc = {
        "Alta/Alb.", "Man./Man.", "Sask./Sask.", "Ont./Ont.",
        "Que/Qc", "N.B./N.-B.", "N.S./N.-É.", "Nfld.Lab./T.-N.-L.",
        "P.E.I./Î.-P.-É.", "Alberta", "Saskatchewan", "Manitoba",
        "Ontario", "Quebec",
    }

    names = set()
    for fn in sorted(os.listdir(cache_dir)):
        if not fn.endswith(".xlsx"):
            continue
        if not any(k in fn.lower() for k in ("completions", "dwelling", "table_2_1")):
            continue

        path = os.path.join(cache_dir, fn)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue

        ws = None
        for sn in ("CSD", "CSD - SDR"):
            if sn in wb.sheetnames:
                ws = wb[sn]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        in_bc = False
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            prov = str(vals[0]).strip() if vals[0] else ""
            if prov in bc_labels:
                in_bc = True
            elif prov in non_bc or (prov.startswith("Canada") and "10,000" in prov):
                in_bc = False
            if not in_bc:
                continue
            csd = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ""
            if "(" in csd and ")" in csd and csd != "Total":
                names.add(csd)
        wb.close()

    return names


# ── Build match list ───────────────────────────────────────────────

def build_matches(pop_entries, cmhc_names):
    rows = []
    for pop_name_full, area_type in pop_entries:
        base = pop_base_name(pop_name_full)
        cmhc = f"{base} ({area_type})"
        if cmhc not in cmhc_names:
            continue

        cmhc_slug = (
            cmhc.replace("(", "").replace(")", "")
            .replace(" ", "_").replace(".", "")
            .strip("_").lower()
        )
        pop_slug = slugify(pop_name_full)

        rows.append({
            "display_name": display_name(base, area_type),
            "cmhc_name": cmhc,
            "pop_name": pop_name_full,
            "cmhc_csv": f"completions_{cmhc_slug}.csv",
            "pop_csv": f"population_{pop_slug}.csv",
            "chart_name": chart_name(base, area_type),
        })

    rows.sort(key=lambda r: r["display_name"])
    return rows


# ── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate all CSVs and mapping.csv for the web frontend."
    )
    parser.add_argument("--pop-cache", default="pop_cache")
    parser.add_argument("--cmhc-cache", default="cmhc_cache")
    parser.add_argument("-o", "--output", default="output")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    python = sys.executable

    # ── Validate caches ────────────────────────────────────────────
    if not os.path.isdir(args.pop_cache):
        sys.exit(f"ERROR: {args.pop_cache}/ not found. Run pop_annual.py first.")
    if not os.path.isdir(args.cmhc_cache):
        sys.exit(f"ERROR: {args.cmhc_cache}/ not found. Run cmhc_annual.py first.")

    # ── Read data ──────────────────────────────────────────────────
    print("Reading population names...")
    pop_entries = read_pop_names(args.pop_cache)
    if pop_entries is None:
        sys.exit("ERROR: Could not read population xlsx from cache.")
    print(f"  {len(pop_entries)} municipalities in BC Stats data.")

    print("Reading CMHC CSD names...")
    cmhc_names = read_cmhc_bc_names(args.cmhc_cache)
    if not cmhc_names:
        sys.exit("ERROR: No BC CSDs found in CMHC cache.")
    print(f"  {len(cmhc_names)} BC CSDs across all cached years.")

    matches = build_matches(pop_entries, cmhc_names)
    print(f"  {len(matches)} municipalities matched.\n")

    if not matches:
        sys.exit("ERROR: No matches found.")

    os.makedirs(args.output, exist_ok=True)

    # ── Generate all CSVs ──────────────────────────────────────────
    cmhc_script = os.path.join(script_dir, "cmhc_annual.py")
    pop_script = os.path.join(script_dir, "pop_annual.py")

    done_cmhc = set()
    done_pop = set()

    for i, m in enumerate(matches, 1):
        label = m["display_name"]
        print(f"[{i}/{len(matches)}] {label}")

        # CMHC completions CSV
        comp_path = os.path.join(args.output, m["cmhc_csv"])
        if m["cmhc_name"] not in done_cmhc and not os.path.isfile(comp_path):
            print(f"  cmhc_annual.py -m {m['cmhc_name']!r}")
            subprocess.run(
                [python, cmhc_script, "-m", m["cmhc_name"], "-o", args.output],
                capture_output=True,
            )
            done_cmhc.add(m["cmhc_name"])
        else:
            print(f"  completions: cached")

        # Population CSV
        pop_path = os.path.join(args.output, m["pop_csv"])
        if m["pop_name"] not in done_pop and not os.path.isfile(pop_path):
            print(f"  pop_annual.py -m {m['pop_name']!r}")
            subprocess.run(
                [python, pop_script, "-m", m["pop_name"], "-o", args.output],
                capture_output=True,
            )
            done_pop.add(m["pop_name"])
        else:
            print(f"  population: cached")

        # Verify
        if os.path.isfile(comp_path) and os.path.isfile(pop_path):
            print(f"  OK")
        else:
            missing = []
            if not os.path.isfile(comp_path):
                missing.append(m["cmhc_csv"])
            if not os.path.isfile(pop_path):
                missing.append(m["pop_csv"])
            print(f"  WARNING: missing {', '.join(missing)}")

    # ── Write mapping.csv ──────────────────────────────────────────
    mapping_path = os.path.join(args.output, "mapping.csv")
    fieldnames = ["display_name", "cmhc_csv", "pop_csv", "chart_name"]
    with open(mapping_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(matches)

    print(f"\nWrote {mapping_path} with {len(matches)} entries.")
    print("Done.")


if __name__ == "__main__":
    main()
